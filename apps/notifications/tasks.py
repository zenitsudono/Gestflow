from celery import shared_task
from django.core.mail import send_mail, EmailMessage
from django.contrib.auth import get_user_model
from django.conf import settings
from django.utils.timezone import now
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ReportLab imports for PDF Generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

User = get_user_model()

@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def send_welcome_email(self, user_id):
    try:
        user = User.objects.get(pk=user_id)
        subject = "Welcome to GestiFlow!"
        message = f"Hello {user.get_full_name()},\n\nWelcome to GestiFlow. Your account has been registered successfully.\n\nBest regards,\nThe GestiFlow Team"
        
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            fail_silently=False,
        )
    except User.DoesNotExist:
        pass
    except Exception as exc:
        raise self.retry(exc=exc)

@shared_task
def send_notification_email(notification_id):
    from .models import Notification
    try:
        notif = Notification.objects.get(pk=notification_id)
        send_mail(
            notif.title,
            notif.message,
            settings.DEFAULT_FROM_EMAIL,
            [notif.recipient.email],
            fail_silently=False,
        )
    except Notification.DoesNotExist:
        pass

@shared_task
def generate_monthly_report(user_id, month, year):
    try:
        user = User.objects.get(pk=user_id)
        from apps.resources.models import Resource
        
        # Aggregate stats for report
        resources = Resource.objects.filter(
            created_at__month=month,
            created_at__year=year
        ).select_related('category', 'owner')
        
        # Build PDF Document
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#1F4E79'),
            spaceAfter=15
        )
        body_style = styles['Normal']
        
        story = []
        story.append(Paragraph(f"GestiFlow Monthly Resource Report ({month}/{year})", title_style))
        story.append(Paragraph(f"Generated for: {user.get_full_name()} ({user.email})", body_style))
        story.append(Paragraph(f"Date: {now().strftime('%B %d, %Y')}", body_style))
        story.append(Spacer(1, 20))
        
        # Table of resources
        table_data = [["Title", "Category", "Status", "Owner", "Department"]]
        for res in resources:
            table_data.append([
                res.title,
                res.category.name,
                res.status.upper(),
                res.owner.get_full_name(),
                res.department or "-"
            ])
            
        t = Table(table_data, colWidths=[150, 100, 80, 100, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAFC')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(t)
        
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Send Email with PDF attachment
        email = EmailMessage(
            f"GestiFlow Monthly Report - {month}/{year}",
            f"Please find attached your monthly resource report for {month}/{year}.",
            settings.DEFAULT_FROM_EMAIL,
            [user.email]
        )
        email.attach(f"Monthly_Report_{month}_{year}.pdf", pdf_data, "application/pdf")
        email.send()
        
    except User.DoesNotExist:
        pass

@shared_task
def export_resources_excel(user_id, filter_params=None):
    try:
        user = User.objects.get(pk=user_id)
        from apps.resources.models import Resource
        
        # Fetch resources
        resources = Resource.objects.filter(is_archived=False).select_related('category', 'owner')
        
        # Apply potential filters
        if filter_params:
            if 'category' in filter_params and filter_params['category']:
                resources = resources.filter(category_id=filter_params['category'])
            if 'status' in filter_params and filter_params['status']:
                resources = resources.filter(status=filter_params['status'])
                
        # Create Excel Sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resources Export"
        
        # Headers
        headers = ["ID", "Title", "Description", "Status", "Category", "Owner Email", "Department", "Created At"]
        ws.append(headers)
        
        # Styling headers
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            
        # Add Data
        for res in resources:
            ws.append([
                str(res.id),
                res.title,
                res.description,
                res.status.upper(),
                res.category.name,
                res.owner.email,
                res.department or "",
                res.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
            
        # Write to memory buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()
        
        # Email export file
        email = EmailMessage(
            "GestiFlow Resources Export",
            "Please find attached your requested resources Excel export file.",
            settings.DEFAULT_FROM_EMAIL,
            [user.email]
        )
        email.attach("resources_export.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        email.send()
        
    except User.DoesNotExist:
        pass
